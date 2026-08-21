# -*- coding: utf-8 -*-
"""Plan B: 쿨메신저 공식 '메시지 다운로드' 기능을 자동화하는 어댑터 (뼈대).

.udb가 SQLite가 아닌 것으로 확인된 경우 이 어댑터를 완성해서 사용합니다.
쿨메신저 자체 UI(메시지 관리함 → 다운로드 아이콘 → 기간 지정 → 다운로드)를
pywinauto로 자동 클릭한 뒤, 기본 저장 경로
(문서\\CoolMessenger Files\\Received Files\\coolmsg_YYYY_MM_DD)에 생성되는
엑셀 파일을 읽어오는 방식입니다.

⚠ 실제 버튼 좌표/컨트롤 이름은 사용자의 쿨메신저 화면을 직접 보며
   pywinauto의 inspect 도구(또는 Claude Code가 실기에서)로 확인해야 완성됩니다.
   여기서는 구조와 엑셀 파싱 부분만 미리 준비해 둡니다.
"""

import glob
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import List

from .base import MessengerAdapter
from ..models import RawMessage


def _download_dir() -> Path:
    return Path.home() / "Documents" / "CoolMessenger Files" / "Received Files"


class CoolMessengerExportAdapter(MessengerAdapter):
    name = "coolmessenger_export"

    def is_available(self) -> bool:
        return _download_dir().exists()

    def trigger_official_download(self, days: int = 1):
        """TODO: pywinauto로 쿨메신저 창을 찾아 다운로드 버튼을 클릭.
        예시 골격:

            from pywinauto import Application
            app = Application(backend="uia").connect(title_re=".*쿨메신저.*")
            win = app.window(title_re=".*쿨메신저.*")
            win["메시지 다운로드"].click_input()
            # 기간 선택 다이얼로그 처리...
            win["다운로드"].click_input()

        실제 컨트롤 이름/구조는 사용자 화면에서 직접 확인 후 채워야 합니다.
        """
        raise NotImplementedError(
            "실제 쿨메신저 UI 컨트롤 구조 확인 후 구현 필요 (README 'Plan B 완성 가이드' 참고)"
        )

    def fetch_recent_messages(self, days: int = 1) -> List[RawMessage]:
        # 1) 공식 다운로드 트리거 (완성 필요)
        # self.trigger_official_download(days)

        # 2) 다운로드된 엑셀 파일 목록 확보
        pattern = str(_download_dir() / "coolmsg_*.xlsx")
        files = sorted(glob.glob(pattern))
        if not files:
            raise FileNotFoundError(
                "다운로드된 쿨메신저 엑셀 파일을 찾지 못했습니다. "
                "trigger_official_download()를 먼저 완성하거나, 수동으로 "
                "'메시지 다운로드'를 한 번 실행해 파일이 생기는지 확인하세요."
            )

        # 3) 엑셀 파싱 (openpyxl 필요 - requirements.txt에 추가해서 사용)
        import openpyxl  # 지연 임포트: Plan A만 쓰는 경우 의존성 불필요

        results: List[RawMessage] = []
        cutoff = datetime.now() - timedelta(days=days)
        latest_file = files[-1]
        wb = openpyxl.load_workbook(latest_file, read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

        def col(name_options):
            for i, h in enumerate(headers):
                if h in name_options:
                    return i
            return None

        idx_sender = col(["보낸사람", "발신자"])
        idx_dept = col(["보낸부서", "부서"])
        idx_title = col(["제목"])
        idx_body = col(["내용", "본문"])
        idx_date = col(["받은시간", "수신일시", "일시"])

        for row_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            try:
                received_at = row[idx_date] if idx_date is not None else datetime.now()
                if isinstance(received_at, str):
                    received_at = datetime.fromisoformat(received_at)
                if received_at < cutoff:
                    continue
                results.append(RawMessage(
                    id=f"{os.path.basename(latest_file)}#{row_i}",
                    sender=str(row[idx_sender]) if idx_sender is not None else "",
                    department=str(row[idx_dept]) if idx_dept is not None else "",
                    title=str(row[idx_title]) if idx_title is not None else "",
                    body=str(row[idx_body]) if idx_body is not None else "",
                    received_at=received_at,
                ))
            except Exception:
                continue

        return results
